from openpyxl import load_workbook

wb = load_workbook(filename='底稿制作.xlsx')
# 处理第一张表单
# sheet_ranges = wb['第一部分 第二章 业务与技术调查']
# 处理第二张表单
sheet_ranges = wb['第一部分 第三章 同业竞争与关联交易调查']
# 新建表单存放格式化后的数据
# new_sheet = wb.create_sheet('formated_第一部分 第二章 业务与技术调查')
new_sheet = wb.create_sheet('formated_第一部分 第三章 同业竞争与关联交易调查')
# 空单元格flag
flag = 0


def indent(uni, content1, content2):
    """根据缩进量，将内容缩进"""
    global flag
    # 起始纯文字，则不移动
    if uni == 0:
        new_sheet[chr(ord('A') + uni) + str(n + 3)].value = content1
        # 对于没有编号的只缩进内容
    if flag > 0:
        # 将excel列号的字母变为数字，加上缩进量再转化回
        new_sheet[chr(ord('B') + uni) + str(n + 3)].value = content2
    else:
        new_sheet[chr(ord('A')+uni)+str(n + 3)].value = content1
        new_sheet[chr(ord('B') + uni) + str(n + 3)].value = content2
    flag = 0
    return


def get_content(n):
    """获取表单第n行的内容"""
    # print(sheet_ranges['A{}'.format(n + 3)])
    content1 = sheet_ranges['A{}'.format(n+3)].value
    content2 = sheet_ranges['B{}'.format(n+3)].value
    return content1, content2


def count_and_indent(n, content):
    """根据‘-’的数量计数缩进量；对于没有编号的行（content[0]类型为None），向前寻找就近的缩进数（回调）"""
    global flag
    # print(n)
    if content[0] != None:
        numb = content[0].count("-")
        # 对于没有编号的只缩进内容,向前查找了flag行，缩进内容时要回到该行
        if flag != 0:
            content = get_content(n+flag)
        indent(numb, content[0], content[1])
    # 起始为空格，向前寻找就近的缩进数
    else:
        n -= 1
        flag += 1
        count_and_indent(n, get_content(n))
    return


# n为表单行数
for n in range(180):
    content = get_content(n)
    count_and_indent(n, content)
wb.save(r'底稿制作.xlsx')